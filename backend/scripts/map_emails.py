"""Extract email mappings from DCLI_ResourceTracker Excel and output SQL."""
import openpyxl
import sys

wb = openpyxl.load_workbook(r"D:\RMS_Siprahub\DCLI_ResourceTracker 1.xlsx", data_only=True)
ws = wb["Resource Data"]
headers = [c.value for c in ws[1]]

name_idx = headers.index("Name")
sipra_idx = headers.index("Siprahub email")
dcli_idx = headers.index("DCLI email")

# Build deduplicated mapping
seen: dict[str, tuple] = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[name_idx]
    sipra = row[sipra_idx]
    dcli = row[dcli_idx]
    if not name:
        continue
    name_str = str(name).strip()
    sipra_str = str(sipra).strip().lower() if sipra and str(sipra).strip().lower() != "none" else None
    dcli_str = str(dcli).strip().lower() if dcli and str(dcli).strip().lower() != "none" else None

    if name_str not in seen:
        seen[name_str] = (sipra_str, dcli_str)
    else:
        old_s, old_d = seen[name_str]
        seen[name_str] = (sipra_str or old_s, dcli_str or old_d)

# Output SQL
sql_parts = []
for name, (sipra, dcli) in seen.items():
    updates = []
    if sipra:
        updates.append(f"siprahub_email = '{sipra}'")
    if dcli:
        updates.append(f"aws_email = '{dcli}'")
    if updates:
        safe_name = name.replace("'", "''")
        set_clause = ", ".join(updates)
        sql_parts.append(
            f"UPDATE employees SET {set_clause} WHERE LOWER(rms_name) = LOWER('{safe_name}');"
        )

print(f"-- Total mappings: {len(sql_parts)}")
for s in sql_parts:
    print(s)
