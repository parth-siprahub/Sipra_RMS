"""
RMS Feature List Generator — v2.0 (Updated March 2, 2026)

Changes from v1.0:
- Phase 1: All 72 features marked Status = "✅ Done"
- Phase 1: 3 new features added (F-073 to F-075)
- Phase 2: 12 original + 8 new features (F-P2-013 to F-P2-020)
- Summary stats updated accordingly
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def make_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_header(cell, bg_color='1F4788'):
    cell.font = Font(bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = make_border()


def apply_data_row(ws, row_num, num_cols, alt=False):
    bg = 'F2F5FB' if alt else 'FFFFFF'
    for col in range(1, num_cols + 1):
        cell = ws.cell(row_num, col)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell.border = make_border()
        cell.alignment = Alignment(wrap_text=True, vertical='top')


def create_updated_feature_list():
    wb = openpyxl.Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ──────────────────────────────────────────────
    # SUMMARY SHEET
    # ──────────────────────────────────────────────
    ws_summary = wb.create_sheet('Summary', 0)

    ws_summary['A1'] = 'RMS Project — Feature List v2.0'
    ws_summary['A1'].font = Font(size=16, bold=True, color='1F4788')
    ws_summary.merge_cells('A1:D1')

    ws_summary['A2'] = 'Updated: March 2, 2026 | Phase 1 Complete ✅ | Demo: Feb 23, 2026'
    ws_summary['A2'].font = Font(size=10, italic=True, color='555555')
    ws_summary.merge_cells('A2:D2')

    info = [
        ('Project Name:', 'Resource Management System (RMS)'),
        ('Document Version:', '2.0 (Updated)'),
        ('Original Version:', '1.0 — February 16, 2026 — Jaicind Santhibhavan'),
        ('Update Date:', 'March 2, 2026'),
        ('', ''),
        ('Phase 1 Timeline:', 'Feb 16 – March 2, 2026 (2 weeks) ✅ DELIVERED'),
        ('Phase 1 Features:', '75 (72 original + 3 added)'),
        ('Phase 1 Status:', '✅ All Done — Demoed to Raja PV & Senthil Natarajan on Feb 23, 2026'),
        ('Phase 1 E2E Tests:', '17 flows verified via Playwright'),
        ('', ''),
        ('Phase 2 Timeline:', 'Post March 2026 (TBD)'),
        ('Phase 2 Features:', '20 (12 original + 8 new)'),
        ('Phase 2 Story Points:', 'TBD'),
        ('', ''),
        ('Tech Stack (Actual):', 'FastAPI + React 18 + TypeScript + Supabase (PostgreSQL)'),
        ('Auth:', 'Supabase Auth (JWT, ES256)'),
        ('Storage:', 'Supabase Storage (resume uploads)'),
        ('DB Tables (Actual):', '6 — profiles, job_profiles, resource_requests, candidates, communication_logs, sows'),
        ('API Endpoints:', '23'),
    ]

    row = 4
    for label, value in info:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True, size=10)
        ws_summary[f'B{row}'].font = Font(size=10)
        row += 1

    ws_summary.column_dimensions['A'].width = 28
    ws_summary.column_dimensions['B'].width = 65

    # ──────────────────────────────────────────────
    # PHASE 1 SHEET
    # ──────────────────────────────────────────────
    ws_phase1 = wb.create_sheet('Phase 1 Features', 1)

    headers = ['Module', 'Feature ID', 'Feature Name', 'Description',
               'Priority', 'Story Points', 'Dependencies', 'Acceptance Criteria', 'Status']
    for col, header in enumerate(headers, 1):
        apply_header(ws_phase1.cell(1, col, header), bg_color='1F4788')

    widths = [22, 14, 38, 52, 12, 12, 22, 62, 14]
    for col, width in enumerate(widths, 1):
        ws_phase1.column_dimensions[get_column_letter(col)].width = width

    ws_phase1.row_dimensions[1].height = 28

    # All 72 original Phase 1 features — Status updated to "✅ Done"
    phase1_features = [
        # ── Authentication & RBAC ──
        ('Authentication & RBAC', 'AUTH-001', 'User Login',
         'Secure login with email and password via Supabase Auth', 'Critical', 8, 'None',
         'User can log in with valid credentials; Invalid login shows error; Session persists across refresh',
         '✅ Done'),
        ('Authentication & RBAC', 'AUTH-002', 'Role-Based Access Control',
         'Three roles: ADMIN, RECRUITER, MANAGEMENT with permission mapping', 'Critical', 12, 'AUTH-001',
         'Recruiter sees recruiter menu; Admin sees admin menu; Management sees dashboards; Permissions enforced at API level',
         '✅ Done'),
        ('Authentication & RBAC', 'AUTH-003', 'Permission-Based Features',
         'Restrict features based on user role', 'High', 8, 'AUTH-002',
         'Recruiter cannot access Send to Client; Admin can process exits; Management has read-only',
         '✅ Done'),
        ('Authentication & RBAC', 'AUTH-004', 'User Management',
         'Admin CRUD for user accounts', 'Medium', 6, 'AUTH-002',
         'Admin creates user; New user can log in; Admin can deactivate users',
         '✅ Done'),

        # ── Dashboard ──
        ('Dashboard & Metrics', 'DASH-001', 'Total Requests Metric',
         'Display total number of resource requests', 'High', 4, 'REQUEST-001',
         'Metric displays correct count; Updates when new request created', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-002', 'Onboarded Count',
         'Display count of onboarded resources', 'High', 3, 'ONBOARD-001',
         'Count matches DB query; Updates when resource onboarded', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-003', 'Awaiting Onboarding Count',
         'Display count of candidates With Client (awaiting onboarding)', 'High', 3, 'ADMIN-006',
         'Shows Status = WITH_CLIENT count; Updates in real-time', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-004', 'To Be Shared Count',
         'Display count of profiles With Admin (to be shared with client)', 'High', 3, 'PIPELINE-013',
         'Shows Status = With Admin count; Admin sees accurate count', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-005', 'Role-Wise Breakdown Chart',
         'Bar/Pie chart showing distribution by role', 'Medium', 8, 'JOBPROF-001',
         'Chart displays; Tooltip shows exact count; Interactive', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-006', 'Technology Distribution Chart',
         'Chart showing distribution by technology', 'Medium', 6, 'JOBPROF-001',
         'Chart groups by technology; Interactive and responsive', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-007', 'Status Filter Dropdown',
         'Filter dashboard by status', 'Medium', 4, 'DASH-001',
         'Selecting status filters dashboard; Metrics update correctly', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-008', 'Attrition Rate Trend',
         'Line graph showing exits over time', 'Low', 8, 'EXIT-001',
         'Line graph shows exits over time; Trend line calculated', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-009', 'Avg Time to Onboard',
         'Metric showing average days from request to onboarding', 'Low', 6, 'ONBOARD-001',
         'Metric calculated correctly; Excludes rejected requests', '✅ Done'),
        ('Dashboard & Metrics', 'DASH-010', 'Date Range Filter',
         'Custom date range selection for dashboard', 'Low', 6, 'DASH-001',
         'Custom date range selection; Accurate filtered data', '✅ Done'),

        # ── Job Profile ──
        ('Job Profile', 'JOBPROF-001', 'Create Job Profile',
         'Create job profile with role, technology, experience', 'High', 8, 'None',
         'Job profile created; Saved to database; Visible in list', '✅ Done'),
        ('Job Profile', 'JOBPROF-002', 'Edit Job Profile',
         'Edit existing job profile', 'High', 4, 'JOBPROF-001',
         'Job profile edited; Changes saved; Updated in list', '✅ Done'),
        ('Job Profile', 'JOBPROF-003', 'Delete Job Profile',
         'Delete job profile with validation (cannot delete if linked)', 'Medium', 4, 'JOBPROF-001',
         'Cannot delete if linked to requests; Confirmation dialog; Deletion successful', '✅ Done'),
        ('Job Profile', 'JOBPROF-004', 'List Job Profiles',
         'Paginated list of job profiles', 'High', 4, 'JOBPROF-001',
         'Profiles listed; Pagination works; Search works', '✅ Done'),
        ('Job Profile', 'JOBPROF-005', 'Duplicate Validation',
         'Prevent duplicate role names', 'Medium', 3, 'JOBPROF-001',
         'Same role name cannot be created twice; Error shown', '✅ Done'),

        # ── Resource Request ──
        ('Resource Request', 'REQUEST-001', 'Create Request',
         'Create resource request with auto-generated ID (REQ-YYYYMMDD-XXX)', 'Critical', 12, 'JOBPROF-001',
         'Request ID auto-generated; Job profile selected; Source captured; Priority set', '✅ Done'),
        ('Resource Request', 'REQUEST-002', 'Job Profile Dropdown',
         'Select job profile from dropdown when creating request', 'High', 4, 'JOBPROF-001',
         'Dropdown populated from job profiles; Selection saved', '✅ Done'),
        ('Resource Request', 'REQUEST-003', 'Request Source',
         'Capture request source (Email/Chat)', 'Medium', 4, 'REQUEST-001',
         'Source dropdown works; Email and Chat options available', '✅ Done'),
        ('Resource Request', 'REQUEST-004', 'Priority Field',
         'Set priority (Urgent/High/Medium/Low)', 'Medium', 3, 'REQUEST-001',
         'Priority dropdown works; All 4 options available', '✅ Done'),
        ('Resource Request', 'REQUEST-005', 'Multi-Position Support',
         'Support requests for multiple positions', 'Medium', 6, 'REQUEST-001',
         'Position count field works; Multiple candidates can be linked', '✅ Done'),
        ('Resource Request', 'REQUEST-006', 'View Request List',
         'Paginated list of all requests', 'High', 8, 'REQUEST-001',
         'Requests listed; Pagination works; Sorting works', '✅ Done'),
        ('Resource Request', 'REQUEST-007', 'Filter Requests',
         'Multi-criteria filtering (status, priority, job profile)', 'Medium', 8, 'REQUEST-006',
         'Filters work independently; Combined filters work; Reset filters works', '✅ Done'),
        ('Resource Request', 'REQUEST-008', 'Search Requests',
         'Search by request ID or candidate name', 'Medium', 6, 'REQUEST-006',
         'Search by ID works; Search by name works; Partial match works', '✅ Done'),

        # ── Recruiter Pipeline ──
        ('Recruiter Pipeline', 'PIPELINE-001', 'Add Candidate (21 Fields)',
         'Add candidate with complete 21-field form', 'Critical', 16, 'REQUEST-001',
         'All 21 fields available; Validation works; Candidate saved', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-002', 'Owner Assignment',
         'Assign recruiter owner from dropdown', 'High', 4, 'AUTH-001',
         'Owner dropdown populated; Selection saved', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-003', 'Vendor Field',
         'Capture sourcing vendor from vendor master', 'Medium', 3, 'PIPELINE-001',
         'Vendor dropdown populated from master; Selection saved', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-004', 'Interview DateTime',
         'Date and time picker for interview', 'Medium', 4, 'PIPELINE-001',
         'Date picker works; Time picker works; Saved correctly', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-005', 'Candidate Status (HR-Approved)',
         'HR-approved status dropdown: NEW, SCREENING_DONE, L1_COMPLETED, L1_REJECTED, L2_COMPLETED, L2_REJECTED, SELECTED, WITH_CLIENT, ONBOARDED, EXIT',
         'High', 6, 'PIPELINE-001',
         'All 10 HR-approved statuses available; Status updates correctly; No legacy statuses',
         '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-006', 'CTC Fields',
         'Current and Expected CTC fields', 'Medium', 3, 'PIPELINE-001',
         'Numeric validation works; Both fields saved', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-007', 'Location Fields',
         'Current and Work Location fields', 'Low', 2, 'PIPELINE-001',
         'Text fields work; Both locations saved', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-008', 'Notice Period',
         'Notice period in days (0–90)', 'Low', 2, 'PIPELINE-001',
         'Numeric field works; Validation works', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-009', 'Remarks Field',
         'Multi-line text for notes (max 1000 chars)', 'Low', 2, 'PIPELINE-001',
         'Textarea works; Long text saved; Character limit enforced', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-010', 'Resume Upload',
         'Upload resume to Supabase Storage (PDF/DOCX, max 5MB)', 'Critical', 8, 'PIPELINE-001',
         'File upload works; PDF and DOCX accepted; 5MB limit enforced; File stored in Supabase',
         '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-011', 'Edit Candidate',
         'Edit all 21 fields of existing candidate', 'High', 6, 'PIPELINE-001',
         'All fields editable; Changes saved; Resume replaceable', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-012', 'View Candidate List',
         'View all candidates for a request', 'High', 8, 'PIPELINE-001',
         'Candidates listed; All fields visible; Sorting works', '✅ Done'),
        ('Recruiter Pipeline', 'PIPELINE-013', 'Mark With Admin',
         'Update request status to With Admin', 'High', 4, 'PIPELINE-001',
         'Status updates; Admin sees request; Request appears in admin queue', '✅ Done'),

        # ── Admin Operations ──
        ('Admin Operations', 'ADMIN-001', 'View With Admin',
         'View all requests marked With Admin', 'High', 6, 'PIPELINE-013',
         'List shows only With Admin requests; Pagination works', '✅ Done'),
        ('Admin Operations', 'ADMIN-002', 'Validation Checklist',
         'Checklist for email, format, JD alignment', 'Medium', 8, 'ADMIN-001',
         'Checklist displays; Items can be checked/unchecked; Validation status saved', '✅ Done'),
        ('Admin Operations', 'ADMIN-003', 'Download Profile',
         'Download candidate profile as PDF', 'High', 4, 'ADMIN-001',
         'PDF generated; Contains all candidate details; Resume included', '✅ Done'),
        ('Admin Operations', 'ADMIN-004', 'Reject to Recruiter',
         'Reject profile back to recruiter with reason', 'Medium', 6, 'ADMIN-001',
         'Rejection reason captured; Status reverts; Recruiter notified', '✅ Done'),
        ('Admin Operations', 'ADMIN-005', 'Generate Email Template',
         'Auto-generate email for client submission', 'High', 8, 'ADMIN-001',
         'Email template generated; Candidate details populated; Editable before send', '✅ Done'),
        ('Admin Operations', 'ADMIN-006', 'Mark With Client',
         'Update status to With Client after submission', 'High', 4, 'ADMIN-005',
         'Status updates; Dashboard metrics update; Request moves to client queue', '✅ Done'),
        ('Admin Operations', 'ADMIN-007', 'Log Communication',
         'Log communication date, client contact, type', 'Medium', 6, 'ADMIN-006',
         'Communication logged; Timestamp saved; Audit trail visible', '✅ Done'),

        # ── Onboarding ──
        ('Onboarding', 'ONBOARD-001', 'Mark Onboarded',
         'Mark candidate as onboarded with billing start date', 'Critical', 8, 'ADMIN-006',
         'Billing date captured; Status updates to ONBOARDED; Dashboard updates', '✅ Done'),
        ('Onboarding', 'ONBOARD-002', 'Client Email',
         'Capture client email ID on onboarding', 'High', 3, 'ONBOARD-001',
         'Email field works; Email validation works; Saved correctly', '✅ Done'),
        ('Onboarding', 'ONBOARD-003', 'Client Jira Username',
         'Capture client Jira username on onboarding', 'High', 3, 'ONBOARD-001',
         'Text field works; Saved correctly; Mapped for billing', '✅ Done'),
        ('Onboarding', 'ONBOARD-004', 'Update to Onboarded',
         'Update candidate status to ONBOARDED', 'High', 4, 'ONBOARD-001',
         'Status updates; Visible in onboarded list; Dashboard updates', '✅ Done'),
        ('Onboarding', 'ONBOARD-005', 'Client Rejection',
         'Handle client rejection with reason', 'Medium', 6, 'ADMIN-006',
         'Rejection reason captured; Status updates; Audit trail maintained', '✅ Done'),
        ('Onboarding', 'ONBOARD-006', 'Replacement Required',
         'Flag if replacement needed on client rejection', 'Medium', 4, 'ONBOARD-005',
         'Checkbox works; Flag saved; Used for backfill creation', '✅ Done'),
        ('Onboarding', 'ONBOARD-007', 'Auto-Create Backfill (Rejection)',
         'Auto-create backfill request on client rejection if replacement needed', 'Medium', 8, 'ONBOARD-006',
         'Backfill request created; Links to original; Inherits job profile; Same SOW carries forward',
         '✅ Done'),

        # ── Exit Management ──
        ('Exit Management', 'EXIT-001', 'Process Exit',
         'Capture exit reason and last working day', 'Critical', 8, 'ONBOARD-001',
         'Exit form works; Reason captured; LWD captured; Status updates to EXIT', '✅ Done'),
        ('Exit Management', 'EXIT-002', 'Exit Reason Dropdown',
         '6 exit reason options: Better Offer, Personal, Performance, Client End, Project End, Other',
         'High', 4, 'EXIT-001',
         'All 6 reasons available; Selection saved', '✅ Done'),
        ('Exit Management', 'EXIT-003', 'Last Working Day',
         'Date picker for last working day', 'High', 3, 'EXIT-001',
         'Date picker works; Saved correctly', '✅ Done'),
        ('Exit Management', 'EXIT-004', 'Exit Notes',
         'Multi-line text for exit notes', 'Low', 2, 'EXIT-001',
         'Textarea works; Long notes saved', '✅ Done'),
        ('Exit Management', 'EXIT-005', 'Replacement Required (Exit)',
         'Flag if replacement needed on exit', 'Medium', 4, 'EXIT-001',
         'Checkbox works; Flag saved; Triggers backfill if Yes', '✅ Done'),
        ('Exit Management', 'EXIT-006', 'Auto-Create Backfill (Exit)',
         'Auto-create backfill on exit if replacement needed; same SOW carries forward', 'Medium', 6, 'EXIT-005',
         'Backfill created; Links to original; Inherits job profile; Same SOW; overlap_until date captured',
         '✅ Done'),
        ('Exit Management', 'EXIT-007', 'Update to Exit',
         'Update candidate status to EXIT', 'High', 3, 'EXIT-001',
         'Status updates; Dashboard updates; Attrition metrics update', '✅ Done'),

        # ── SOW Tracker ──
        ('SOW Tracker', 'SOW-001', 'Manual SOW Entry',
         'Form to manually enter SOW details (Admin only)', 'Medium', 6, 'None',
         'SOW form works; All fields saved; Validation works', '✅ Done'),
        ('SOW Tracker', 'SOW-002', 'Link SOW to Requests',
         'SOW is mandatory-linked to each resource request', 'Medium', 6, 'SOW-001',
         'SOW selection required on request creation; Association saved; Validated at API level',
         '✅ Done'),
        ('SOW Tracker', 'SOW-003', 'View SOW List (Active Default)',
         'Table view of all SOWs — default filter = Active SOWs only', 'Medium', 6, 'SOW-001',
         'Active SOWs shown by default; Filter to show inactive/all works; Pagination works',
         '✅ Done'),
        ('SOW Tracker', 'SOW-004', 'Request Count per SOW',
         'Display count of requests and onboarded resources per SOW', 'Low', 4, 'SOW-002',
         'Linked request count displayed; Onboarded instances count correct; Updates on change',
         '✅ Done'),

        # ── Infrastructure ──
        ('Infrastructure', 'INFRA-001', 'Supabase Setup',
         'Create and configure Supabase project with PostgreSQL + Auth + Storage', 'Critical', 4, 'None',
         'Project created; Database provisioned; Auth configured; Storage bucket active', '✅ Done'),
        ('Infrastructure', 'INFRA-002', 'Database Schema (6 Tables)',
         'Design and create 6 tables: profiles, job_profiles, resource_requests, candidates, communication_logs, sows',
         'Critical', 16, 'INFRA-001',
         'All 6 tables created; Relationships defined; Constraints applied; RLS policies set',
         '✅ Done'),
        ('Infrastructure', 'INFRA-003', 'FastAPI Backend (Async)',
         'Scaffold FastAPI backend with async architecture (23 endpoints)', 'Critical', 8, 'INFRA-002',
         'FastAPI running on Uvicorn; Async operations working; Supabase connected; 23 endpoints live',
         '✅ Done'),
        ('Infrastructure', 'INFRA-004', 'React Frontend',
         'Setup React 18 + TypeScript + Vite frontend', 'Critical', 6, 'None',
         'React app running; Routing works; API communication works; TypeScript types clean',
         '✅ Done'),
        ('Infrastructure', 'INFRA-005', 'API Endpoints (23)',
         'Define and implement 23 REST API endpoints', 'High', 8, 'INFRA-003',
         'All endpoints documented; CRUD operations work; Error handling implemented; Auth guards active',
         '✅ Done'),
        ('Infrastructure', 'INFRA-006', 'CORS Configuration',
         'Configure CORS for frontend-backend communication', 'Medium', 2, 'INFRA-003',
         'CORS allows frontend domain; No CORS errors in browser', '✅ Done'),
        ('Infrastructure', 'INFRA-007', 'Environment Variables',
         'Setup .env for secrets management', 'Medium', 3, 'INFRA-003',
         'Env vars loaded; Secrets secured; No hardcoded credentials', '✅ Done'),

        # ── NEW FEATURES (added during development — not in v1.0) ──
        ('Vendor Management', 'F-073', 'Vendor Management Module',
         'CRUD module for vendor master — WRS, GFM, Internal, Anton vendors. Admin-only write access. Vendor master drives dropdown across candidate forms and dashboard analytics.',
         'High', 8, 'AUTH-002, PIPELINE-003',
         'Admin can add/edit/deactivate vendors; Vendor dropdown in candidate form pulls from master; Vendor sourcing analytics reflected in dashboard',
         '✅ Done'),
        ('Recruiter Pipeline', 'F-074', 'Kanban View — Candidate Pipeline',
         'Visual Kanban board grouping candidates by pipeline status column. Supports drag-to-update status. Provides at-a-glance view of all candidates across stages.',
         'High', 10, 'PIPELINE-001, PIPELINE-005',
         'Kanban board renders all pipeline status columns; Candidates appear in correct column; Column counts accurate; Responsive layout',
         '✅ Done'),
        ('Recruiter Pipeline', 'F-075', 'L1/L2 Interview Feedback Capture',
         'Dedicated fields on candidate record to capture L1 and L2 round-specific interview feedback and outcomes. Feeds into status transitions (L1_COMPLETED, L1_REJECTED, L2_COMPLETED, L2_REJECTED).',
         'High', 6, 'PIPELINE-001, PIPELINE-005',
         'L1 feedback field available; L2 feedback field available; Feedback saved on candidate record; Status correctly transitions to L1/L2 outcomes',
         '✅ Done'),
    ]

    row = 2
    for idx, feature_tuple in enumerate(phase1_features):
        (module, feat_id, name, desc, priority, points, deps, ac, status) = feature_tuple
        ws_phase1.cell(row, 1, module)
        ws_phase1.cell(row, 2, feat_id)
        ws_phase1.cell(row, 3, name)
        ws_phase1.cell(row, 4, desc)
        ws_phase1.cell(row, 5, priority)
        ws_phase1.cell(row, 6, points)
        ws_phase1.cell(row, 7, deps)
        ws_phase1.cell(row, 8, ac)
        ws_phase1.cell(row, 9, status)

        apply_data_row(ws_phase1, row, 9, alt=(idx % 2 == 0))

        # colour the status cell green
        status_cell = ws_phase1.cell(row, 9)
        status_cell.fill = PatternFill(start_color='D6F4D6', end_color='D6F4D6', fill_type='solid')
        status_cell.font = Font(bold=True, color='1A7A1A', size=10)
        status_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # priority colour
        priority_cell = ws_phase1.cell(row, 5)
        priority_colors = {
            'Critical': ('FFE5E5', 'CC0000'),
            'High': ('FFF3CD', '856404'),
            'Medium': ('D1ECF1', '0C5460'),
            'Low': ('E2E3E5', '383D41'),
        }
        if priority in priority_colors:
            bg, fg = priority_colors[priority]
            priority_cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            priority_cell.font = Font(bold=True, color=fg, size=10)
            priority_cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1

    # Freeze header
    ws_phase1.freeze_panes = 'A2'

    # ──────────────────────────────────────────────
    # PHASE 2 SHEET
    # ──────────────────────────────────────────────
    ws_phase2 = wb.create_sheet('Phase 2 Features', 2)

    for col, header in enumerate(headers, 1):
        apply_header(ws_phase2.cell(1, col, header), bg_color='2E5C8A')

    for col, width in enumerate(widths, 1):
        ws_phase2.column_dimensions[get_column_letter(col)].width = width

    ws_phase2.row_dimensions[1].height = 28

    phase2_features = [
        # ── Original 12 Phase 2 features ──
        ('Client Portal', 'PORTAL-001', 'Client Login',
         'Direct login for clients to view their request profiles', 'High', 'TBD', 'AUTH-001',
         'Client can log in; Sees only their requests; Cannot see other clients data',
         'Phase 2 — Planned'),
        ('Client Portal', 'PORTAL-002', 'Profile View & Feedback',
         'Client can view candidate profiles and provide accept/reject feedback', 'High', 'TBD', 'PORTAL-001',
         'Profile details visible; Feedback form works; Admin notified',
         'Phase 2 — Planned'),

        ('Advanced Analytics', 'ANALYTICS-001', 'Predictive Attrition',
         'ML-based model to predict attrition risk for onboarded resources', 'Medium', 'TBD', 'EXIT-001',
         'Model trained; Risk score displayed; Recommendations shown',
         'Phase 2 — Planned'),
        ('Advanced Analytics', 'ANALYTICS-002', 'Custom Report Builder',
         'Drag-and-drop custom report builder with export to PDF', 'Low', 'TBD', 'DASH-001',
         'Builder UI works; Custom reports generated; Export works',
         'Phase 2 — Planned'),

        ('Integrations', 'INTEG-001', 'Email Integration',
         'Gmail/Outlook sync for automated request creation from emails', 'Medium', 'TBD', 'REQUEST-001',
         'Emails synced; Requests auto-created; Attachments extracted',
         'Phase 2 — Planned'),
        ('Integrations', 'INTEG-002', 'Calendar Integration',
         'Sync interview schedules to Google/Outlook calendar', 'Medium', 'TBD', 'PIPELINE-004',
         'Calendar events created; Reminders sent; Two-way sync works',
         'Phase 2 — Planned'),
        ('Integrations', 'INTEG-003', 'Jira API Integration',
         'Sync onboarded resources to client Jira (pending client approval for API access)', 'High', 'TBD', 'ONBOARD-003',
         'Jira account auto-created; Username synced; Tickets created',
         'Phase 2 — Planned'),

        ('Mobile App', 'MOBILE-001', 'iOS App',
         'Native iOS app for recruiters and admins', 'Low', 'TBD', 'AUTH-001',
         'iOS app works; Login works; Core features accessible',
         'Phase 2 — Planned'),
        ('Mobile App', 'MOBILE-002', 'Android App',
         'Native Android app for recruiters and admins', 'Low', 'TBD', 'AUTH-001',
         'Android app works; Login works; Core features accessible',
         'Phase 2 — Planned'),

        ('Advanced Features', 'ADV-001', 'Bulk Upload',
         'Excel import for candidate batch onboarding', 'Medium', 'TBD', 'PIPELINE-001',
         'Excel template works; Bulk import works; Validation on import',
         'Phase 2 — Planned'),
        ('Advanced Features', 'ADV-002', 'Document Version Control',
         'Version history for resume uploads', 'Low', 'TBD', 'PIPELINE-010',
         'Versions tracked; Can revert to previous; Diff visible',
         'Phase 2 — Planned'),
        ('Advanced Features', 'ADV-003', 'Multi-Language Support',
         'UI in multiple languages (English, Hindi)', 'Low', 'TBD', 'None',
         'Language switcher works; All UI translated; Saves preference',
         'Phase 2 — Planned'),

        # ── New Phase 2 features (v2.0 additions) ──
        ('Payroll Management', 'F-P2-013', 'Payroll Management — Jira Dump Import',
         'Import Jira time tracking CSV exports and calculate payroll for onboarded resources. Client will not allow direct Jira API access — HR team provides monthly CSV dumps. System calculates billable hours × rate and generates payroll records.',
         'High', 'TBD', 'ONBOARD-001',
         'CSV upload works; Jira hours mapped to correct resources; Payroll calculation correct; Summary report generated; Export to Excel/PDF',
         'Phase 2 — Planned'),

        ('Exit Management', 'F-P2-014', 'Exit Management Full Flow (Dedicated Page)',
         'Dedicated Exit Management page with full exit workflow: initiation, reason capture, LWD, replacement decision, overlap tracking, and exit analytics. Phase 1 implements basic exit within candidate record; this is a dedicated full-flow page.',
         'High', 'TBD', 'EXIT-001',
         'Dedicated exit page renders; All exit fields available; Replacement workflow complete; Exit analytics dashboard shows attrition trends',
         'Phase 2 — Planned'),

        ('Onboarding', 'F-P2-015', 'Onboarding Workflow Page',
         'Dedicated Onboarding Workflow page for structured onboarding: billing date capture, client Jira ID assignment, onboarding checklist, and documentation tracking. Extends basic onboarding fields from Phase 1.',
         'High', 'TBD', 'ONBOARD-001',
         'Dedicated onboarding page renders; Checklist available; Billing date and Jira ID captured; Onboarding status tracked; Notification sent on completion',
         'Phase 2 — Planned'),

        ('Infrastructure', 'F-P2-016', 'Production Hosting & Deployment',
         'Deploy the full RMS stack to a production cloud environment (Railway/Fly.io or equivalent). Includes CI/CD pipeline, HTTPS, custom domain, and environment separation (dev/prod).',
         'Critical', 'TBD', 'INFRA-001',
         'App accessible via HTTPS on production URL; Zero-downtime deployment; CI/CD pipeline runs on push; Prod DB separate from dev',
         'Phase 2 — Planned'),

        ('Infrastructure', 'F-P2-017', 'PR Review Workflow Setup',
         'Configure GitHub branch protection rules and PR review workflow per Jaicind\'s process: daily commits → PR creation → manager approval → merge to main. Includes required reviews and status check gates.',
         'High', 'TBD', 'INFRA-004',
         'Branch protection enabled on main; PRs require at least 1 reviewer approval; Status checks must pass; Direct push to main blocked; Workflow documented',
         'Phase 2 — Planned'),

        ('Documentation & Training', 'F-P2-018', 'User Training Materials',
         'Create user-facing training materials: SOPs for Back Office Recruiters and Admins, video walkthroughs of key flows, quick reference cards, and FAQ document.',
         'High', 'TBD', 'None',
         'Recruiter SOP document complete; Admin SOP document complete; At least 3 key workflow videos recorded; Quick reference cards printed; FAQ answers 20+ questions',
         'Phase 2 — Planned'),

        ('AI Features', 'F-P2-019', 'Resume Parsing (AI-Powered)',
         'AI-powered extraction of candidate details from uploaded PDF/DOCX resumes. Auto-populate candidate form fields (name, email, skills, experience, current company) from resume content to reduce manual entry.',
         'Medium', 'TBD', 'PIPELINE-010',
         'Resume parsed on upload; Name, email, skills auto-populated; Accuracy ≥ 80% on standard formats; User can override any populated field',
         'Phase 2 — Planned'),

        ('Integrations', 'F-P2-020', 'LinkedIn / Job Board API Integration',
         'Integration with LinkedIn and/or third-party job boards for direct candidate sourcing into the RMS pipeline. Mentioned by Senthil Natarajan in stakeholder call. Reduces time spent on manual candidate entry from job boards.',
         'Medium', 'TBD', 'PIPELINE-001',
         'LinkedIn profile import works; Job board search integrated; Candidate pre-filled from public profile; Duplicate detection works',
         'Phase 2 — Planned'),
    ]

    row = 2
    for idx, feature_tuple in enumerate(phase2_features):
        (module, feat_id, name, desc, priority, points, deps, ac, status) = feature_tuple
        ws_phase2.cell(row, 1, module)
        ws_phase2.cell(row, 2, feat_id)
        ws_phase2.cell(row, 3, name)
        ws_phase2.cell(row, 4, desc)
        ws_phase2.cell(row, 5, priority)
        ws_phase2.cell(row, 6, points)
        ws_phase2.cell(row, 7, deps)
        ws_phase2.cell(row, 8, ac)
        ws_phase2.cell(row, 9, status)

        apply_data_row(ws_phase2, row, 9, alt=(idx % 2 == 0))

        # Highlight new v2.0 additions in light blue
        if feat_id.startswith('F-P2-0'):
            for col in range(1, 10):
                cell = ws_phase2.cell(row, col)
                cell.fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')

        row += 1

    ws_phase2.freeze_panes = 'A2'

    # ──────────────────────────────────────────────
    # CHANGELOG SHEET
    # ──────────────────────────────────────────────
    ws_log = wb.create_sheet('Change Log', 3)

    ws_log['A1'] = 'Feature List — Change Log'
    ws_log['A1'].font = Font(size=14, bold=True, color='1F4788')
    ws_log.merge_cells('A1:D1')

    log_headers = ['Date', 'Change', 'Details', 'Author']
    for col, header in enumerate(log_headers, 1):
        apply_header(ws_log.cell(2, col, header), bg_color='1F4788')

    log_entries = [
        ('Feb 16, 2026', 'v1.0 Created',
         'Initial feature list: 72 Phase 1 features, 12 Phase 2 features', 'Jaicind Santhibhavan'),
        ('Mar 2, 2026', 'v2.0 Update — Phase 1 Status',
         'All 72 Phase 1 features marked ✅ Done. Phase 1 delivered on time. Demo to Raja PV and Senthil Natarajan on Feb 23, 2026.',
         'Parth P'),
        ('Mar 2, 2026', 'v2.0 Update — New P1 Features',
         'Added F-073 (Vendor Management Module), F-074 (Kanban View), F-075 (L1/L2 Interview Feedback) — built but not in original list.',
         'Parth P'),
        ('Mar 2, 2026', 'v2.0 Update — Phase 2 Expanded',
         'Added 8 new Phase 2 items: F-P2-013 (Payroll/Jira), F-P2-014 (Exit Full Flow), F-P2-015 (Onboarding Workflow), F-P2-016 (Production Hosting), F-P2-017 (PR Review), F-P2-018 (Training), F-P2-019 (Resume AI), F-P2-020 (LinkedIn API). New items highlighted in blue.',
         'Parth P'),
        ('Mar 2, 2026', 'v2.0 Update — Candidate Statuses',
         'Updated PIPELINE-005 to reflect 10 HR-approved statuses: NEW, SCREENING_DONE, L1_COMPLETED, L1_REJECTED, L2_COMPLETED, L2_REJECTED, SELECTED, WITH_CLIENT, ONBOARDED, EXIT (supersedes legacy v1.0 statuses).',
         'Parth P'),
        ('Mar 2, 2026', 'v2.0 Update — Infrastructure',
         'Updated INFRA-002 description: actual DB is 6 tables (not 11 as projected in v1.0). Updated INFRA-003/005 to reflect actual 23 endpoints.',
         'Parth P'),
    ]

    for row_offset, (date, change, details, author) in enumerate(log_entries):
        r = 3 + row_offset
        ws_log.cell(r, 1, date)
        ws_log.cell(r, 2, change)
        ws_log.cell(r, 3, details)
        ws_log.cell(r, 4, author)
        for col in range(1, 5):
            cell = ws_log.cell(r, col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = make_border()
            if row_offset % 2 == 0:
                cell.fill = PatternFill(start_color='F2F5FB', end_color='F2F5FB', fill_type='solid')

    ws_log.column_dimensions['A'].width = 16
    ws_log.column_dimensions['B'].width = 30
    ws_log.column_dimensions['C'].width = 70
    ws_log.column_dimensions['D'].width = 22
    ws_log.row_dimensions[2].height = 20

    # ──────────────────────────────────────────────
    # SAVE
    # ──────────────────────────────────────────────
    import os
    output_dir = r'd:\RMS_Siprahub\PRD_Feature'
    output_path = os.path.join(output_dir, 'RMS_Feature_List_v2_Updated.xlsx')
    wb.save(output_path)

    print('✅ RMS_Feature_List_v2_Updated.xlsx created successfully!')
    print(f'   Location: {output_path}')
    print(f'   Phase 1: 75 features (72 original ✅ Done + 3 new)')
    print(f'   Phase 2: 20 features (12 original + 8 new — highlighted blue)')
    print(f'   Change Log sheet included')
    print(f'   Ready for stakeholder distribution')


if __name__ == '__main__':
    create_updated_feature_list()
