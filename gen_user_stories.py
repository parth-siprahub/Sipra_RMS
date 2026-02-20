"""
User Stories Excel Generator — Extracted from PRD
10 user stories with full acceptance criteria, persona, epic mapping
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_user_stories():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'User Stories'
    
    headers = ['Story ID', 'Title', 'As a...', 'I want to...', 'So that...', 'Acceptance Criteria', 'Priority', 'Epic', 'Sprint', 'Status']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    widths = [12, 30, 22, 45, 40, 70, 12, 25, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    stories = [
        (
            'US-001',
            'Resource Request Creation',
            'Back Office Recruiter',
            'create resource requests with auto-generated IDs',
            'I can track client requirements systematically',
            '• Request ID auto-generated (REQ-YYYYMMDD-XXX)\n• Job profile selected from dropdown\n• Request source (Email/Chat) captured\n• Priority (Urgent/High/Medium/Low) can be set\n• Multi-position requests supported\n• Request saved and visible in list view',
            'Critical',
            'Resource Request Workflow',
            'Sprint 2',
            'To Do',
        ),
        (
            'US-002',
            'Candidate Addition (21-Field Form)',
            'Back Office Recruiter',
            'add candidates with complete profiles',
            'Admin can review all necessary information',
            '• All 21 fields fillable (see PRD Section 4)\n• Resume uploadable (PDF/DOCX, max 5MB)\n• Owner (recruiter) assigned from dropdown\n• Status defaults to "New"\n• Candidate linked to request ID\n• Form validates required fields before save',
            'Critical',
            'Recruiter Pipeline',
            'Sprint 3',
            'To Do',
        ),
        (
            'US-003',
            'Admin Profile Review',
            'Admin',
            'review recruiter-submitted profiles',
            'I only send quality candidates to clients',
            '• View all requests "With Admin"\n• See complete candidate profile with resume\n• Reject to recruiter with reason\n• Approve and proceed to client submission\n• Validation checklist visible (email, JD match, resume quality)\n• Track L1/L2 interview response status',
            'Critical',
            'Admin Operations',
            'Sprint 4',
            'To Do',
        ),
        (
            'US-004',
            'Client Submission',
            'Admin',
            'generate professional emails for client submission',
            'the process is standardized and auditable',
            '• Email template auto-generated with candidate details\n• Can download candidate profile as PDF\n• Can log communication date and client contact\n• Status updates to "With Client" after submission\n• Audit trail maintained\n• Profile rejection flow handled (backfill yes/no)',
            'Critical',
            'Admin Operations',
            'Sprint 4',
            'To Do',
        ),
        (
            'US-005',
            'Onboarding Workflow',
            'Admin',
            'capture onboarding details',
            'billing and client systems are synchronized',
            '• Mark candidate as onboarded\n• Billing start date captured\n• Client email ID captured\n• Client Jira username captured\n• Status updates to "Onboarded"\n• Dashboard metrics update in real-time',
            'Critical',
            'Lifecycle Management',
            'Sprint 5',
            'To Do',
        ),
        (
            'US-006',
            'Exit Processing',
            'Admin',
            'process exits with reasons',
            'attrition trends can be analyzed',
            '• Select exit reason from dropdown (6 options: Better Offer, Personal, Performance, Client End, Project End, Other)\n• Last working day captured\n• Exit notes can be added\n• Replacement required flag can be set\n• Backfill request auto-created if replacement needed\n• Status updates to "Exit"',
            'Critical',
            'Lifecycle Management',
            'Sprint 5',
            'To Do',
        ),
        (
            'US-007',
            'Dashboard Visibility',
            'Management / Leadership',
            'see real-time metrics',
            'I can track pipeline health and make informed decisions',
            '• Total requests count displayed\n• Onboarded count displayed\n• Awaiting onboarding count (status = "With Client") shown\n• To be shared count (status = "With Admin") shown\n• Role-wise breakdown chart displayed\n• Technology distribution visible\n• Resource status visible at each stage',
            'High',
            'Dashboard & Metrics',
            'Sprint 6',
            'To Do',
        ),
        (
            'US-008',
            'SOW Tracker',
            'Admin',
            'track SOW details',
            'headcount can be reconciled with contracts',
            '• Manually enter SOW details\n• Link request IDs to SOW\n• View SOW list in table format (default: active SOWs)\n• See request count per SOW\n• Active SOW count = number of active resources\n• Filter by SOW status (Active/Expired/All)',
            'Medium',
            'Lifecycle Management',
            'Sprint 5',
            'To Do',
        ),
        (
            'US-009',
            'Backfill Automation',
            'Admin',
            'have backfill requests auto-created',
            'replacements are not forgotten when exits or rejections occur',
            '• Backfill created when exit has "replacement required" = Yes\n• Backfill created when client rejection has "replacement required" = Yes\n• Backfill request inherits job profile from original\n• Backfill linked to original request ID',
            'High',
            'Lifecycle Management',
            'Sprint 5',
            'To Do',
        ),
        (
            'US-010',
            'Communication Logging',
            'Admin',
            'log all client communications',
            'there is a complete audit trail for compliance',
            '• Log communication date\n• Log client contact person\n• Log communication type (email/call/meeting)\n• All logs timestamped and linked to request',
            'Medium',
            'Admin Operations',
            'Sprint 4',
            'To Do',
        ),
    ]

    for row_idx, story in enumerate(stories, 2):
        for col_idx, val in enumerate(story, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='Calibri', size=11)

    # ============== SHEET 2: AC Matrix ==============
    ws2 = wb.create_sheet('Acceptance Criteria Matrix')
    
    ac_headers = ['Story ID', 'Criteria #', 'Acceptance Criterion', 'Testable?', 'Automation Priority']
    for col, h in enumerate(ac_headers, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    ac_widths = [12, 12, 60, 12, 18]
    for col, w in enumerate(ac_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    row = 2
    for story in stories:
        story_id = story[0]
        criteria = story[5].split('\n')
        for i, crit in enumerate(criteria, 1):
            crit_text = crit.strip().lstrip('•').strip()
            if crit_text:
                ws2.cell(row, 1, story_id).font = Font(name='Calibri', size=11)
                ws2.cell(row, 2, f'AC-{i:02d}').font = Font(name='Calibri', size=11)
                ws2.cell(row, 3, crit_text).font = Font(name='Calibri', size=11)
                ws2.cell(row, 3).alignment = Alignment(wrap_text=True)
                ws2.cell(row, 4, 'Yes').font = Font(name='Calibri', size=11)
                ws2.cell(row, 5, 'E2E').font = Font(name='Calibri', size=11)
                row += 1

    output = r'C:/Users/parth/.gemini/antigravity/brain/39b25fbd-59cc-4ae8-8351-8dd232f82e33/RMS_User_Stories.xlsx'
    wb.save(output)
    print(f'User Stories Excel generated: {output}')
    print(f'  Sheet 1: 10 user stories with full AC')
    print(f'  Sheet 2: AC matrix ({row - 2} testable criteria)')

if __name__ == '__main__':
    create_user_stories()
