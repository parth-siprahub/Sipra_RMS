import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_feedback_sheet():
    file_path = r"d:\RMS_Siprahub\docs\SipraHub_User_Feedback.xlsx"
    
    # Define columns
    columns = [
        "Component",
        "Issue/Gap Description",
        "Feedback Type",
        "Severity/Priority",
        "Reported By",
        "Date Reported",
        "Status",
        "Admin Comments"
    ]
    
    # Create an empty DataFrame
    df = pd.DataFrame(columns=columns)
    
    # Create Excel writer
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='User Feedback')
        workbook = writer.book
        worksheet = writer.sheets['User Feedback']
        
        # Define dropdown options
        components = ['Dashboard', 'SOW', 'Job Profiles', 'Candidates', 'Vendors', 'Auth', 'Search', 'Other']
        feedback_types = ['Bug', 'Feature Request', 'Enhancement', 'UX/UI', 'Performance']
        priorities = ['Critical (P0)', 'High (P1)', 'Medium (P2)', 'Low (P3)']
        statuses = ['New', 'Triaged', 'In Progress', 'Resolved', "Won't Fix"]
        
        # Create Data Validations
        dv_component = DataValidation(type="list", formula1=f'"{",".join(components)}"', allow_blank=True)
        dv_type = DataValidation(type="list", formula1=f'"{",".join(feedback_types)}"', allow_blank=True)
        dv_priority = DataValidation(type="list", formula1=f'"{",".join(priorities)}"', allow_blank=True)
        dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
        
        # Add error messages
        dv_component.error ='Your entry is not in the list'
        dv_component.errorTitle = 'Invalid Entry'
        
        # Target ranges (e.g., A2:A100)
        worksheet.add_data_validation(dv_component)
        dv_component.add('A2:A100')
        
        worksheet.add_data_validation(dv_type)
        dv_type.add('C2:C100')
        
        worksheet.add_data_validation(dv_priority)
        dv_priority.add('D2:D100')
        
        worksheet.add_data_validation(dv_status)
        dv_status.add('G2:G100')
        
        # Formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Dark Green
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border
            
            # Set column width
            column_letter = cell.column_letter
            if column_title == 'Issue/Gap Description':
                worksheet.column_dimensions[column_letter].width = 50
            elif column_title == 'Admin Comments':
                worksheet.column_dimensions[column_letter].width = 40
            else:
                worksheet.column_dimensions[column_letter].width = 20

    print(f"Excel feedback sheet created at: {file_path}")

if __name__ == "__main__":
    create_feedback_sheet()
